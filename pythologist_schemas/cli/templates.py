#!/usr/bin/env python

""" Generate templates of inputs for pythologist.

Inputs can vary depending on the platform.
"""

import re, sys
from datetime import datetime
import argparse, gzip
import json
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl import Workbook
from importlib_resources import files

highlight = NamedStyle(name="highlight")

def cli():
   args = do_inputs()
   main(args)

def main(args):
   "We need to take the platform and return an appropriate input template"

   # Lets start with the panel
   if args.panel_output: do_panel_output(args)

   return

def do_panel_output(args):
   #import schema_data.inputs as schema_data_inputs

   #from .schemas import inputs as schemas_inputs
   _schema = json.loads(files('schema_data.inputs').joinpath('panel.json').read_text())

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)

   _oname = args.panel_output

   # Start with the Metadata. Write the header and the value names

   ws1 = wb.create_sheet("Meta")
   header_names = ['Metadata Field','Metadata Value']
   for _j,_header_name in enumerate(header_names):
       ws1.cell(row=1,column=_j+1).style = highlight
       ws1.cell(row=1,column=_j+1).value = _header_name
    
   for _i, _property in enumerate(_schema['properties']['meta']['properties']):
       ws1.cell(row=_i+2,column=1).value=_property

   # Now lets make the Panel.  Write the header only.
   ws2 = wb.create_sheet("Panel")
   header_names = list(_schema['properties']['markers']['items']['properties'])
   #print(header_names)
   for _j,_header_name in enumerate(header_names):
      entry = _schema['properties']['markers']['items']['properties'][_header_name]
      print(entry)
      ws2.cell(row=1,column=_j+1).style = highlight
      ws2.cell(row=1,column=_j+1).value = _header_name if 'title' not in entry else entry['title']    
    
   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = _oname)

def do_inputs():
   parser = argparse.ArgumentParser(
            description = "",
            formatter_class=argparse.ArgumentDefaultsHelpFormatter)
   parser.add_argument('--platform',metavar='platform_name',choices=['InForm','InForm plus'],default='InForm',help="The input platform format")
   parser.add_argument('--panel_output',help="How to output the panel variable")
   args = parser.parse_args()
   return args

def external_cmd(cmd):
   """function for calling program by command through a function"""
   cache_argv = sys.argv
   sys.argv = cmd
   args = do_inputs()
   main(args)
   sys.argv = cache_argv


if __name__ == "__main__":
	main(do_inputs())