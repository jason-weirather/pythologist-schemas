#!/usr/bin/env python

""" Generate templates of inputs for pythologist.

Inputs can vary depending on the platform.
"""

import re, sys
from datetime import datetime
import argparse, gzip
import json
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from importlib_resources import files
from pythologist_schemas import get_validator

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)
highlight.fill = PatternFill(start_color="CDEAC2", end_color="CDEAC2", fill_type = "solid")

boldened = NamedStyle(name="boldened")
boldened.font = Font(bold=True)

def cli():
   args = do_inputs()
   main(args)

def main(args):
   "We need to take the platform and return an appropriate input template"

   # Lets start with the panel

   #if args.panel_output: do_panel_output(args)
   if args.project_output: 
      do_project_folder_output(args.project_output)
   if args.analysis_output: 
      do_analysis_output(args.analysis_output)
   if args.report_output: 
      do_report_output(args.report_output)

   return

def _write_parameters(worksheet,fields):
   "Write the metadata fields to the worksheet"
   "fields can be either a single object of metadata, or multiple in list or tuple to stack"
   header_names = ['Parameter','Value']
   for _j,_header_name in enumerate(header_names):
      worksheet.cell(row=1,column=_j+1).style = highlight
      worksheet.cell(row=1,column=_j+1).value = _header_name 
   
   is_list = False
   if type(fields) is list:
      is_list = True
   elif type(fields) is tuple:
      is_list = True

   # If we got a list of fields consoldate their properties
   if is_list:
      _f = {}
      _f['properties'] = {}
      for _field in fields:
         for _k in _field['properties']:
            _f['properties'][_k] = _field['properties'][_k]
      fields = _f

   for _i, _property in enumerate(fields['properties']):
      worksheet.cell(row=_i+2,column=1).style = boldened
      worksheet.cell(row=_i+2,column=1).value=  fields['properties'][_property]['title']
      if 'default' in fields['properties'][_property]:
         worksheet.cell(row=_i+2,column=2).value = fields['properties'][_property]['default']
         
         

def _write_repeating(worksheet,fields):
   "Write the repeating data fields to the worksheet"
   header_names = list(fields['items']['properties'])
   #print(header_names)
   for _j,_header_name in enumerate(header_names):
      entry = fields['items']['properties'][_header_name]
      print(entry)
      worksheet.cell(row=1,column=_j+1).style = highlight
      worksheet.cell(row=1,column=_j+1).value = _header_name if 'title' not in entry else entry['title']

def _fix_width(worksheet,min_width=20,padding=3):
   column_widths = []
   for row in worksheet:
      for i, cell in enumerate(row):
         if cell.value is None: continue
         if len(column_widths) > i:
            if len(str(cell.value)) > column_widths[i]:
                column_widths[i] = min_width if len(str(cell.value))+padding < min_width else len(str(cell.value))+padding
         else:
            column_widths += [min_width if len(str(cell.value))+padding < min_width else len(str(cell.value))+padding]
   for i, column_width in enumerate(column_widths):
      worksheet.column_dimensions[get_column_letter(i+1)].width = column_width


def do_report_output(output_path):
   _validator = get_validator(files('schema_data.inputs').joinpath('report_definition.json'))
   _schema = _validator.schema
   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)


   # Start with the Metadata. Write the header and the value names

   ws0 = wb.create_sheet(_schema['properties']['parameters']['title'])
   _write_parameters(ws0,_schema['properties']['parameters'])
   _fix_width(ws0)

   ws1 = wb.create_sheet(_schema['properties']['population_percentages']['title'])
   _write_repeating(ws1,_schema['properties']['population_percentages'])
   _fix_width(ws1)

   ws2 = wb.create_sheet(_schema['properties']['population_densities']['title'])
   _write_repeating(ws2,_schema['properties']['population_densities'])
   _fix_width(ws2)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = output_path)
   return

def do_analysis_output(output_file):
   _validator1 = get_validator(files('schema_data.inputs').joinpath('panel.json'))
   _validator2 = get_validator(files('schema_data.inputs.platforms.InForm').joinpath('analysis.json'))
   _schema1 = _validator1.schema
   _schema2 = _validator2.schema

   #_schema1 = json.loads(files('schema_data.inputs').joinpath('panel.json').read_text())
   #_schema2 = json.loads(files('schema_data.inputs.platforms.InForm').joinpath('analysis.json').read_text())

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)


   # Start with the Metadata. Write the header and the value names

   ws0 = wb.create_sheet(_schema2['properties']['parameters']['title'])
   _write_parameters(ws0,[_schema1['properties']['parameters'],_schema2['properties']['parameters']])
   _fix_width(ws0)

   ws1 = wb.create_sheet(_schema1['properties']['markers']['title'])
   _write_repeating(ws1,_schema1['properties']['markers'])
   _fix_width(ws1)

   ws2 = wb.create_sheet(_schema2['properties']['inform_exports']['title'])
   _write_repeating(ws2,_schema2['properties']['inform_exports'])
   _fix_width(ws2)

   ws3 = wb.create_sheet(_schema2['properties']['mutually_exclusive_phenotypes']['title'])
   _write_repeating(ws3,_schema2['properties']['mutually_exclusive_phenotypes'])
   _fix_width(ws3)

   ws4 = wb.create_sheet(_schema2['properties']['binary_phenotypes']['title'])
   _write_repeating(ws4,_schema2['properties']['binary_phenotypes'])
   _fix_width(ws4)

   ws5 = wb.create_sheet(_schema2['properties']['regions']['title'])
   _write_repeating(ws5,_schema2['properties']['regions'])
   _fix_width(ws5)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = output_file)
   return

def do_project_folder_output(output_file):
   # For now lets keep this with InForm only
   _validator = get_validator(files('schema_data.inputs.platforms.InForm').joinpath('project.json'))
   _schema = _validator.schema

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)

   # Start with the Metadata. Write the header and the value names

   ws1 = wb.create_sheet(_schema['properties']['parameters']['title'])
   _write_parameters(ws1,_schema['properties']['parameters'])
   _fix_width(ws1)


   # Now lets make the Panel.  Write the header only.
   ws2 = wb.create_sheet(_schema['properties']['samples']['title'])
   _write_repeating(ws2,_schema['properties']['samples'])
   _fix_width(ws2)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = output_file)

def do_panel_output(args):
   #import schema_data.inputs as schema_data_inputs

   _validator = get_validator(files('schema_data.inputs').joinpath('panel.json'))
   _schema = _validator.schema

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)

   _oname = args.panel_output

   # Start with the Metadata. Write the header and the value names

   ws1 = wb.create_sheet(_schema['properties']['parameters']['title'])
   _write_parameters(ws1,_schema['properties']['parameters'])
   _fix_width(ws1)


   # Now lets make the Panel.  Write the header only.
   ws2 = wb.create_sheet(_schema['properties']['markers']['title'])
   _write_repeating(ws2,_schema['properties']['markers'])
   _fix_width(ws2)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = _oname)

def do_inputs():
   parser = argparse.ArgumentParser(
            description = "",
            formatter_class=argparse.ArgumentDefaultsHelpFormatter)
   parser.add_argument('--platform',metavar='platform_name',choices=['InForm'],default='InForm',help="The input platform format")
   #parser.add_argument('--panel_output',help="How to output the panel variable") # no need to output panel only since for InForm we wrap it into analysis
   parser.add_argument('--project_output',help="How to output the panel variable")
   parser.add_argument('--analysis_output',help="How to output the panel variable")
   parser.add_argument('--report_output',help="How to output the panel variable")
   args = parser.parse_args()
   return args

def external_cmd(cmd):
   """function for calling program by command through a function"""
   cache_argv = sys.argv
   sys.argv = cmd
   args = do_inputs()
   main(args)
   sys.argv = cache_argv


if __name__ == "__main__":
	main(do_inputs())